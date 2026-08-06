"""Runs every dataset in a config file through the census_transform steps."""

import glob
import json
import os

from openpyxl import load_workbook

from . import census_transform, io_utils, proportions, recipes


# Reads the dataset list from a config file, resolving relative raw file paths against base_path
def load_dataset_config(config_path: str, base_path: str = "") -> list:
    with open(os.path.expanduser(config_path), "r") as f:
        config = json.load(f)

    if base_path:
        base_path = os.path.expanduser(base_path)
        for dataset in config["datasets"]:
            if not os.path.isabs(dataset["original_file_path"]):
                dataset["original_file_path"] = os.path.join(base_path, dataset["original_file_path"])

    return config["datasets"]


# Copies one dataset's raw file to the centralized folder, then maps aliases and drops columns
def process_census_dataset(key: str, alias: str, original_file_path: str, file_blocks: dict, central_path_head: str):
    centralized_file_dir = os.path.join(central_path_head, alias)
    file_blocks[key]["original_file_path"] = original_file_path
    file_blocks[key]["centralized_file_dir"] = centralized_file_dir

    file_name = os.path.basename(original_file_path)
    file_path = os.path.join(centralized_file_dir, file_name)

    io_utils.copy_file(original_file_path, centralized_file_dir)

    cols_to_drop = file_blocks[key]["cols_to_drop"]
    column_mapping_dict = file_blocks[key]["code_to_alias_column_mappings"]

    print(f"Processing {alias}")

    alias_df = census_transform.map_census_aliases(file_path, column_mapping_dict)
    processed_df = census_transform.census_drop_cols(alias_df, cols_to_drop)

    return processed_df


# Runs every dataset listed in config_path through process_census_dataset, keyed by alias
def load_and_process_all_datasets(config_path: str, base_path: str, file_blocks: dict, central_path_head: str) -> dict:
    dataset_configs = load_dataset_config(config_path, base_path)

    processed_datasets = {}

    for config in dataset_configs:
        try:
            df = process_census_dataset(
                key=config["key"],
                alias=config["alias"],
                original_file_path=config["original_file_path"],
                file_blocks=file_blocks,
                central_path_head=central_path_head,
            )
            processed_datasets[config["alias"]] = df

        except Exception as e:
            print(f"Error processing {config['alias']}: {e}")
        print("-" * 30)

    return processed_datasets


# Reads JPL notes spreadsheet into one block per bold row.
# A column only gets kept if "subfield to keep" cell has a value.
def build_file_blocks(notes_wb_path: str) -> dict:
    notes_wb = load_workbook(os.path.expanduser(notes_wb_path))
    sheet = notes_wb.active

    file_blocks = {}
    current_file_name = None
    cols_not_to_drop = ["Geography", "Geographic Area Name"]

    for row in sheet.iter_rows(min_row=1, max_col=3, values_only=False):
        cell_value = row[0].value
        is_bold = row[0].font.bold if row[0].font else False

        if is_bold and cell_value:
            original_name = cell_value.strip()
            current_file_name = io_utils.to_snake_case(original_name)
            file_blocks[current_file_name] = {
                "original_name": original_name,
                "cols_to_drop": [],
                "code_to_alias_column_mappings": {},
                "original_file_path": "",
                "centralized_file_dir": "",
            }
            continue

        if current_file_name and any(col.value for col in row):
            col_name = row[0].value
            subfield_value = row[1].value
            if col_name and not subfield_value and col_name not in cols_not_to_drop:
                cleaned_col_name = col_name.strip().strip("'\"")
                cleaned_col_name = " ".join(cleaned_col_name.split())
                file_blocks[current_file_name]["cols_to_drop"].append(cleaned_col_name)

    return file_blocks


"""
Runs every dataset from raw CSV to its final cleaned:
Returns {output_name: cleaned_df}
If output_dir is given, also writes each df out as a CSV.
"""
def run_full_cleaning(
    config_path: str,
    base_path: str,
    notes_wb_path: str,
    central_path_head: str,
    output_dir: str = None,
) -> dict:
    file_blocks = build_file_blocks(notes_wb_path)
    datasets = load_and_process_all_datasets(config_path, base_path, file_blocks, central_path_head)

    cleaned = {
        "age_of_structure": recipes.recipe_age_of_structure(datasets["age_of_structure"]),
        "aggregate_vehicles": recipes.merge_tenure_households(datasets["aggregate_vehicles"], datasets["tenure"]),
        "health_insurance": recipes.recipe_health_insurance(datasets["health_insurance"]),
        "households_w_computer": datasets["households_w_computer"],
        "internet_subscription": datasets["internet_subscription"],
        "limited_english_speaking": recipes.recipe_limited_english_speaking(datasets["limited_english_speaking"]),
        "living_arrangements": recipes.recipe_living_arrangements(datasets["living_arrangements"]),
        "population_group_quarters": datasets["population_group_quarters"],
        "race_origin": datasets["race_origin"],
        "tenure": datasets["tenure"],
    }

    person_outputs = recipes.recipe_person_under_5_65(datasets["person_under_5_65"])
    cleaned["genders"] = person_outputs["genders"]
    cleaned["person_under_5_65_males"] = person_outputs["males"]
    cleaned["person_under_5_65_females"] = person_outputs["females"]

    # Hawaiian Homelands also carries poverty-status columns that belong on the FPL
    # dataset, so the raw (pre-recipe) df is kept around for that merge below.
    hawaiian_homelands_raw_df = datasets["2022_census_hawaiian_homelands"]
    cleaned["2022_census_hawaiian_homelands"] = recipes.recipe_hawaiian_homelands(hawaiian_homelands_raw_df)

    fpl_df = recipes.recipe_income_share_of_fpl(datasets["income_share_of_fpl"])
    cleaned["income_share_of_fpl"] = recipes.merge_hawaiian_homelands_poverty(fpl_df, hawaiian_homelands_raw_df)

    if output_dir:
        for name, df in cleaned.items():
            census_transform.export_census_csv(df, output_dir, f"{name}.csv", overwrite=True)

    return cleaned


# person_under_5_65 split outputs share one pop total but have no config alias of their own
SPLIT_OUTPUT_DENOMINATORS = {
    "genders": "Estimate!!Total:",
    "person_under_5_65_males": "Estimate!!Total:",
    "person_under_5_65_females": "Estimate!!Total:",
}


# Adds Census_Population and "(%)" columns to every CSV in cleaned_dir, in place
def add_percentages_to_directory(
    cleaned_dir: str,
    config_path: str,
    base_path: str,
    block_group_populations: dict,
    hawaiian_homelands_populations: dict,
) -> None:
    cleaned_dir = os.path.expanduser(cleaned_dir)
    total_population = sum(block_group_populations.values()) + sum(hawaiian_homelands_populations.values())

    dataset_configs = load_dataset_config(config_path, base_path)
    denominator_map = proportions.build_denominator_map(dataset_configs)

    for csv_file in glob.glob(os.path.join(cleaned_dir, "*.csv")):
        alias = os.path.splitext(os.path.basename(csv_file))[0]
        denominator_column = denominator_map.get(alias) or SPLIT_OUTPUT_DENOMINATORS.get(alias)
        print(f"Adding proportions to {os.path.basename(csv_file)}...")
        df_with_props = proportions.add_percentages_to_csv(
            csv_file,
            total_population,
            block_group_populations,
            hawaiian_homelands_populations,
            denominator_column=denominator_column,
        )
        df_with_props.to_csv(csv_file, index=False)
